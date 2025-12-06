import openpyxl


class ExcelUtility:

    def __init__(self, filename, sheet):
        self.file_name = filename
        self.workbook = openpyxl.load_workbook(filename=self.file_name)
        self.worksheet = self.workbook[sheet]

    def get_row_count(self):
        return self.worksheet.max_row

    def get_column_count(self):
        return self.worksheet.max_column

    def read_data_from_cell(self,row,col):
        return self.worksheet.cell(row=row, column=col).value

    def write_data_to_cell(self,row,col,data):
        self.worksheet.cell(row=row, column=col).value = data
        self.workbook.save(self.file_name)
